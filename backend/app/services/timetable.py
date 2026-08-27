"""课表解析与归一化。

支持来源：Wisedu（金智教务）接口返回的行数据、Excel、HTML 表格、ICS 日历。
统一输出 ``CourseDraft`` 结构，供「预览确认后保存」与「生成计划」使用。
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import Optional

from .. import schemas


# 默认作息：新疆政法学院（目前唯一内置学校）。其它学校可在设置中手动调整。
DEFAULT_PERIOD_TIMES: list[dict[str, str]] = [
    {"start": "10:00", "end": "10:45"},
    {"start": "10:50", "end": "11:35"},
    {"start": "11:50", "end": "12:30"},
    {"start": "12:35", "end": "13:25"},
    {"start": "13:30", "end": "14:15"},
    {"start": "16:00", "end": "16:45"},
    {"start": "16:55", "end": "17:35"},
    {"start": "17:50", "end": "18:35"},
    {"start": "18:40", "end": "19:25"},
    {"start": "20:30", "end": "21:15"},
    {"start": "21:20", "end": "22:05"},
]


def default_period_times() -> list[dict[str, str]]:
    return [dict(p) for p in DEFAULT_PERIOD_TIMES]


def infer_term(today: Optional[date] = None) -> str:
    """按当前日期推断学年学期，例如 2025-2026-1（秋季）或 2025-2026-2（春季）。"""
    today = today or date.today()
    y, m = today.year, today.month
    if m >= 9:
        return f"{y}-{y + 1}-1"
    if m == 1:
        return f"{y - 1}-{y}-1"
    return f"{y - 1}-{y}-2"


def mask_to_weeks(mask: str) -> list[int]:
    """把形如 1111111100000000 的周次掩码转成上课周列表（第 N 位为 1 即第 N 周上课）。"""
    weeks = [i + 1 for i, ch in enumerate(mask or "") if ch == "1"]
    return weeks


def weeks_to_mask(weeks: list[int], length: int = 20) -> str:
    chars = ["0"] * length
    for w in weeks:
        if 1 <= w <= length:
            chars[w - 1] = "1"
    return "".join(chars)


def parse_week_label(label: Optional[str]) -> list[int]:
    """解析中文周次描述，如 1-8周、1-16周(单)、2-16周(双)、1-8,10-12周。"""
    text = (label or "").replace(" ", "")
    if not text:
        return []
    odd = "单" in text
    even = "双" in text
    weeks: set[int] = set()
    for m in re.finditer(r"(\d+)\s*[-–~]\s*(\d+)", text):
        start, end = int(m.group(1)), int(m.group(2))
        for w in range(start, end + 1):
            if odd and w % 2 == 0:
                continue
            if even and w % 2 == 1:
                continue
            weeks.add(w)
    tokens = [t for t in re.split(r"[第周,\s、]+", text) if t and t.isdigit()]
    for t in tokens:
        n = int(t)
        if odd and n % 2 == 0:
            continue
        if even and n % 2 == 1:
            continue
        weeks.add(n)
    return sorted(weeks)


def resolve_weeks(week_mask: Optional[str], week_label: Optional[str]) -> list[int]:
    """优先使用周次掩码，缺失时从周次描述解析。"""
    if week_mask and re.fullmatch(r"[01]{8,32}", week_mask):
        weeks = mask_to_weeks(week_mask)
        if weeks:
            return weeks
    return parse_week_label(week_label)


def normalize_day(value) -> int:
    """把 星期一 / 周一 / 4 / 四 等转成 1-7。"""
    if value is None:
        return 0
    s = str(value).strip()
    if s.isdigit():
        d = int(s)
        return d if 1 <= d <= 7 else 0
    mapping = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7}
    for ch in s:
        if ch in mapping:
            return mapping[ch]
    return 0


def parse_period(value) -> tuple[int, int]:
    """把 1-2节 / 第1-2节 / 1,2 等转成 (start, end)。"""
    if value is None:
        return (1, 1)
    s = str(value).strip()
    m = re.search(r"(\d+)\s*[-–~,、]\s*(\d+)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d+)", s)
    if m:
        n = int(m.group(1))
        return n, n
    return (1, 1)


def _clean(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_float(value) -> Optional[float]:
    s = _clean(value)
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _norm(value: Optional[str]) -> str:
    return (value or "").strip().lower().replace(" ", "").replace("：", "")


COLUMN_ALIASES: dict[str, list[str]] = {
    "name": ["课程名称", "课程名", "名称", "课程", "课名", "科目", "kcm", "课程科目"],
    "code": ["课程代码", "课号", "课程编号", "kch"],
    "teacher": ["教师", "任课教师", "授课教师", "老师", "主讲教师", "skjs", "教师姓名"],
    "location": ["教室", "上课地点", "地点", "教学地点", "场所", "jasmc", "上课教室", "教室名称"],
    "day": ["星期", "周几", "上课日", "星期几", "xq", "skxq", "星期名称"],
    "period": ["节次", "上课节次", "节序号", "节", "时间", "ksjc", "jsjc", "第几节"],
    "week": ["周次", "上课周次", "教学周次", "周数", "skzc", "zcmc"],
    "credit": ["学分", "xf", "总学分"],
    "type": ["课程性质", "类型", "课程类型", "kcxz", "课程类别"],
}


def _map_headers(cells: list[Optional[str]]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(cells):
        n = _norm(cell)
        if not n:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if field in mapping:
                continue
            for alias in aliases:
                a = _norm(alias)
                if a and (n == a or (len(a) >= 2 and a in n)):
                    mapping[field] = idx
                    break
            if field in mapping:
                break
    return mapping


def _cell(cells: list, idx: Optional[int]):
    if idx is None or idx >= len(cells):
        return None
    return cells[idx]


def _draft(**kwargs) -> dict:
    return {
        "term": kwargs["term"],
        "name": kwargs["name"],
        "code": kwargs.get("code") or None,
        "teacher": kwargs.get("teacher") or None,
        "location": kwargs.get("location") or None,
        "day_of_week": kwargs["day_of_week"],
        "start_period": kwargs["start_period"],
        "end_period": kwargs["end_period"],
        "week_mask": kwargs.get("week_mask") or None,
        "week_label": kwargs.get("week_label") or None,
        "credit": kwargs.get("credit") or None,
        "course_type": kwargs.get("course_type") or None,
    }


def parse_wisedu_rows(rows: list[dict], term: Optional[str] = None, warnings: Optional[list[str]] = None) -> list[dict]:
    """把金智教务接口返回的 rows 归一化为 CourseDraft 字典。"""
    warnings = warnings if warnings is not None else []
    courses: list[dict] = []
    for row in rows or []:
        name = _clean(row.get("KCM"))
        if not name:
            warnings.append("跳过一条缺少课程名称的记录")
            continue
        day = int(row.get("SKXQ") or 0)
        if day < 1 or day > 7:
            warnings.append(f"跳过“{name}”：星期字段无效")
            continue
        raw_term = _clean(row.get("XNXQDM")) or term or infer_term()
        week_mask = _clean(row.get("SKZC"))
        if week_mask and not re.fullmatch(r"[01]{8,32}", week_mask):
            week_mask = None
        courses.append(
            _draft(
                term=raw_term,
                name=name,
                code=_clean(row.get("KCH")),
                teacher=_clean(row.get("SKJS")),
                location=_clean(row.get("JASMC")),
                day_of_week=day,
                start_period=int(row.get("KSJC") or 1),
                end_period=int(row.get("JSJC") or row.get("KSJC") or 1),
                week_mask=week_mask,
                week_label=_clean(row.get("ZCMC")),
                credit=_to_float(row.get("XF")),
                course_type=_clean(row.get("KCXZDM_DISPLAY") or row.get("KCLBDM_DISPLAY")),
            )
        )
    return courses


def _parse_teacher_week(text: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    """从 like '9-16周[实践]/张德慧[主讲]' 解析出教师和周次。"""
    teacher: Optional[str] = None
    week_label: Optional[str] = None
    for part in (text or "").split("/"):
        part = part.strip()
        if not part:
            continue
        head = part.split("[")[0].strip()
        if "周" in head:
            week_label = head
        elif "[" in part and head:
            teacher = head
    return teacher, week_label


def parse_arranged_list(items: list[dict], term: Optional[str] = None, warnings: Optional[list[str]] = None) -> list[dict]:
    """解析新教务接口 getMyScheduleDetail 的 datas.arrangedList 为 CourseDraft。"""
    warnings = warnings if warnings is not None else []
    courses: list[dict] = []
    for item in items or []:
        name = _clean(item.get("courseName"))
        if not name:
            warnings.append("跳过一条缺少课程名称的记录")
            continue
        day = int(item.get("dayOfWeek") or 0)
        if day < 1 or day > 7:
            warnings.append(f"跳过“{name}”：星期字段无效")
            continue
        teacher, week_label = _parse_teacher_week(_clean(item.get("weeksAndTeachers")))
        week_mask = _clean(item.get("week"))
        if week_mask and not re.fullmatch(r"[01]{8,32}", week_mask):
            week_mask = None
        courses.append(
            _draft(
                term=term or infer_term(),
                name=name,
                code=_clean(item.get("courseCode")),
                teacher=teacher,
                location=_clean(item.get("placeName")),
                day_of_week=day,
                start_period=int(item.get("beginSection") or 1),
                end_period=int(item.get("endSection") or item.get("beginSection") or 1),
                week_mask=week_mask,
                week_label=week_label,
                credit=_to_float(item.get("credit")),
                course_type=None,
            )
        )
    return courses


def _parse_flat_rows(rows: list[list], term: Optional[str], warnings: list[str]) -> list[dict]:
    header_idx: Optional[int] = None
    header_map: dict[str, int] = {}
    for i, row in enumerate(rows[:30]):
        cells = [_clean(c) for c in row]
        mapping = _map_headers(cells)
        if mapping.get("name") is not None and (mapping.get("day") is not None or mapping.get("period") is not None):
            header_idx = i
            header_map = mapping
            break
    if header_idx is None:
        warnings.append("未能识别表格表头，请确认文件为课程表格式")
        return []

    courses: list[dict] = []
    for row in rows[header_idx + 1 :]:
        cells = row
        if all(not _clean(c) for c in cells):
            continue
        name = _clean(_cell(cells, header_map.get("name")))
        if not name:
            continue
        day = normalize_day(_cell(cells, header_map.get("day")))
        if day == 0:
            warnings.append(f"无法识别“{name}”的上课星期，已跳过")
            continue
        start_p, end_p = parse_period(_cell(cells, header_map.get("period")))
        week_raw = _clean(_cell(cells, header_map.get("week")))
        week_mask = None
        week_label = week_raw
        if week_raw and re.fullmatch(r"[01]{8,32}", week_raw):
            week_mask = week_raw
        else:
            weeks = parse_week_label(week_raw)
            week_mask = weeks_to_mask(weeks) if weeks else None
        courses.append(
            _draft(
                term=term or infer_term(),
                name=name,
                code=_clean(_cell(cells, header_map.get("code"))),
                teacher=_clean(_cell(cells, header_map.get("teacher"))),
                location=_clean(_cell(cells, header_map.get("location"))),
                day_of_week=day,
                start_period=start_p,
                end_period=end_p,
                week_mask=week_mask,
                week_label=week_label,
                credit=_to_float(_cell(cells, header_map.get("credit"))),
                course_type=_clean(_cell(cells, header_map.get("type"))),
            )
        )
    return courses


def parse_excel(data: bytes, term: Optional[str] = None, warnings: Optional[list[str]] = None) -> list[dict]:
    from openpyxl import load_workbook

    warnings = warnings if warnings is not None else []
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return _parse_flat_rows(rows, term, warnings)


def _cell_text(node) -> Optional[str]:
    from bs4 import BeautifulSoup, NavigableString

    text = node.get_text(" ", strip=True) if hasattr(node, "get_text") else str(node)
    return _clean(text)


def _parse_html_table(table, term: Optional[str], warnings: list[str]) -> list[dict]:
    rows_raw = []
    for tr in table.find_all("tr"):
        cells = [_cell_text(c) for c in tr.find_all(["td", "th"])]
        rows_raw.append(cells)
    if not rows_raw:
        return []
    courses = _parse_flat_rows(rows_raw, term, warnings)
    if courses:
        return courses
    # 尝试网格排布：表头为星期，行首为节次
    return _parse_grid_rows(rows_raw, term, warnings)


def _parse_grid_rows(rows: list[list[Optional[str]]], term: Optional[str], warnings: list[str]) -> list[dict]:
    if len(rows) < 2:
        return []
    header = rows[0]
    days: list[int] = []
    for cell in header[1:]:
        d = normalize_day(cell)
        days.append(d)
    if not any(days):
        return []
    courses: list[dict] = []
    for row in rows[1:]:
        period_raw = row[0] if row else None
        start_p, end_p = parse_period(period_raw)
        for idx, day in enumerate(days):
            if day == 0 or idx + 1 >= len(row):
                continue
            block = row[idx + 1]
            if not block:
                continue
            name, teacher, location, week_label = _classify_grid_block(block)
            if not name:
                warnings.append("网格中存在无法解析出课程名的格子，已跳过")
                continue
            weeks = parse_week_label(week_label)
            courses.append(
                _draft(
                    term=term or infer_term(),
                    name=name,
                    teacher=teacher or None,
                    location=location or None,
                    day_of_week=day,
                    start_period=start_p,
                    end_period=end_p,
                    week_mask=weeks_to_mask(weeks) if weeks else None,
                    week_label=week_label or None,
                )
            )
    return courses


def _classify_grid_block(block: str) -> tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    lines = [l.strip() for l in re.split(r"[\n\r]+", block) if l.strip()]
    name: Optional[str] = None
    teacher: Optional[str] = None
    location: Optional[str] = None
    week_label: Optional[str] = None
    for line in lines:
        if "周" in line and not name:
            week_label = line
            continue
        if ("教师" in line or "老师" in line) and "教室" not in line:
            teacher = re.sub(r"^(教师|老师|授课教师|任课教师)[:：]?", "", line).strip()
            if teacher:
                continue
        if any(k in line for k in ("教室", "楼", "馆", "区", "地址", "房间")):
            location = re.sub(r"^(教室|上课地点|地点)[:：]?", "", line).strip()
            if location:
                continue
        if name is None:
            name = line
    if name and "(单)" in name:
        week_label = week_label or "单周"
    return name, teacher, location, week_label


def parse_html(text: str, term: Optional[str] = None, warnings: Optional[list[str]] = None) -> list[dict]:
    from bs4 import BeautifulSoup

    warnings = warnings if warnings is not None else []
    soup = BeautifulSoup(text, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        warnings.append("未在 HTML 中找到表格")
        return []
    courses: list[dict] = []
    for table in tables:
        courses.extend(_parse_html_table(table, term, warnings))
    seen = set()
    unique: list[dict] = []
    for c in courses:
        key = (c["term"], c["day_of_week"], c["start_period"], c["end_period"], c["name"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(c)
    return unique


def parse_ics(text: str, term: Optional[str] = None, warnings: Optional[list[str]] = None) -> list[dict]:
    from icalendar import Calendar

    warnings = warnings if warnings is not None else []
    try:
        cal = Calendar.from_ical(text.encode("utf-8"))
    except Exception:
        warnings.append("ICS 解析失败，请确认文件为有效 iCalendar 日历")
        return []
    courses: list[dict] = []
    for component in cal.walk("VEVENT"):
        summary = _clean(component.get("SUMMARY"))
        if not summary:
            continue
        name = summary
        teacher: Optional[str] = None
        if "@" in summary:
            name, teacher = summary.rsplit("@", 1)
        dtstart = component.get("DTSTART")
        dtend = component.get("DTEND")
        if not dtstart or not getattr(dtstart, "dt", None):
            continue
        start_dt = dtstart.dt
        end_dt = dtend.dt if dtend and getattr(dtend, "dt", None) else None
        day = start_dt.isoweekday()
        start_fmt = start_dt.strftime("%H:%M")
        end_fmt = end_dt.strftime("%H:%M")
        start_p, end_p = _periods_from_time(start_fmt, end_fmt)
        weeks = _weeks_from_rrule(component, start_dt)
        courses.append(
            _draft(
                term=term or infer_term(),
                name=name.strip(),
                teacher=teacher or None,
                location=_clean(component.get("LOCATION")),
                day_of_week=day,
                start_period=start_p,
                end_period=end_p,
                week_mask=weeks_to_mask(weeks) if weeks else None,
                week_label=(f"{weeks[0]}-{weeks[-1]}周" if weeks else None),
            )
        )
    return courses


def _periods_from_time(start_fmt: str, end_fmt: str) -> tuple[int, int]:
    start_p, end_p = 1, 1
    for idx, p in enumerate(DEFAULT_PERIOD_TIMES, start=1):
        if p["start"] == start_fmt:
            start_p = idx
        if p["end"] == end_fmt:
            end_p = idx
    if end_p < start_p:
        end_p = start_p
    return start_p, end_p


def _weeks_from_rrule(component, start_dt) -> list[int]:
    rrule = component.get("RRULE")
    count = 1
    if rrule and hasattr(rrule, "get"):
        count_val = rrule.get("COUNT")
        if count_val:
            try:
                count = int(count_val[0])
            except (TypeError, ValueError):
                count = 1
    return list(range(1, count + 1))


def parse_payload(
    file_data: Optional[bytes],
    filename: Optional[str],
    raw_text: Optional[str],
    term: Optional[str],
) -> schemas.ParseTimetableOut:
    warnings: list[str] = []
    courses: list[dict] = []
    detected_term = term

    source = (raw_text or "").strip()
    fname = (filename or "").lower()
    if source:
        if "BEGIN:VCALENDAR" in source or fname.endswith(".ics"):
            parsed = parse_ics(source, term, warnings)
            courses.extend(parsed)
            detected_term = detected_term or infer_term()
        elif "<table" in source or "<tr" in source or "<html" in source or fname.endswith((".html", ".htm")):
            parsed = parse_html(source, term, warnings)
            courses.extend(parsed)
            detected_term = detected_term or infer_term()
        else:
            # 粘贴的纯文本：按行/制表符或逗号拆成表格
            rows = [line.split("\t") for line in source.splitlines() if line.strip()]
            parsed = _parse_flat_rows(rows, term, warnings)
            courses.extend(parsed)
            detected_term = detected_term or infer_term()
    elif file_data:
        if fname.endswith((".xlsx", ".xls")):
            try:
                courses = parse_excel(file_data, term, warnings)
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"Excel 解析失败：{exc}")
        elif fname.endswith(".ics"):
            try:
                courses = parse_ics(file_data.decode("utf-8", errors="replace"), term, warnings)
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"ICS 解析失败：{exc}")
        else:
            text = file_data.decode("utf-8", errors="replace")
            if "<table" in text or "<tr" in text:
                courses = parse_html(text, term, warnings)
            else:
                rows = [line.split("\t") for line in text.splitlines() if line.strip()]
                courses = _parse_flat_rows(rows, term, warnings)
        detected_term = detected_term or infer_term()
    else:
        warnings.append("未提供任何课表数据")

    terms = {c["term"] for c in courses if c.get("term")}
    if len(terms) == 1:
        detected_term = terms.pop()

    drafts = [schemas.CourseDraft(**c) for c in courses]
    return schemas.ParseTimetableOut(term=detected_term, courses=drafts, warnings=warnings)
